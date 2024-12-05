import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# List of CSV file names
csv_files = [
    "file1.csv",
    "file2.csv",
    "file3.csv",
    "file4.csv",
    "file5.csv",
    "file6.csv",
    "file7.csv",
    "file8.csv",
]

# Excel file name to save
output_excel = "merged_excel.xlsx"

# Create a new workbook
wb = Workbook()
ws = wb.active

# Initialize starting row
current_row = 1

# Loop through each file and process
for i, file in enumerate(csv_files):
    # Insert an empty row above jksm501 onwards (after the first CSV)
    if i > 0:
        ws.insert_rows(current_row)

    # Add a row with the custom name (e.g., jksm500) merged across 3 cells and bold
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    cell = ws.cell(row=current_row, column=1)
    cell.value = f"jksm500{i}"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    current_row += 1

    # Read the CSV file, preserving 'N/A'
    df = pd.read_csv(file, keep_default_na=False, na_values=["N/A"])

    # Add the CSV index (column headers) in bold
    for col_index, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=current_row, column=col_index)
        cell.value = column_name
        cell.font = Font(bold=True)
    current_row += 1

    # Add the CSV data below the index
    for _, row in df.iterrows():
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=current_row, column=col_index, value=value)
        current_row += 1

    # Add an empty row after the CSV data
    current_row += 1  # This ensures there is a gap between datasets

# Save the workbook
wb.save(output_excel)

print(f"Merged Excel file has been saved as '{output_excel}'")

###
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# List of CSV file names
csv_files = [
    "file1.csv",
    "file2.csv",
    "file3.csv",
    "file4.csv",
    "file5.csv",
    "file6.csv",
    "file7.csv",
    "file8.csv",
]

# Excel file name to save
output_excel = "merged_excel.xlsx"

# Create a new workbook
wb = Workbook()
ws = wb.active

# Initialize starting row
current_row = 1

# Loop through each file and process
for i, file in enumerate(csv_files):
    # Insert an empty row above jksm501 onwards (after the first CSV)
    if i > 0:
        ws.insert_rows(current_row)

        # Merge the first 3 cells in the newly inserted row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"jksm500{i}"  # Custom label for the merged cells
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        current_row += 1  # Move to the next row after the merged cells

    # Add a row with the custom name (e.g., jksm500) merged across 3 cells and bold
    if i == 0:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"jksm500{i}"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        current_row += 1

    # Read the CSV file, preserving 'N/A'
    df = pd.read_csv(file, keep_default_na=False, na_values=["N/A"])

    # Add the CSV index (column headers) in bold
    for col_index, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=current_row, column=col_index)
        cell.value = column_name
        cell.font = Font(bold=True)
    current_row += 1

    # Add the CSV data below the index
    for _, row in df.iterrows():
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=current_row, column=col_index, value=value)
        current_row += 1

    # Add an empty row after the CSV data
    current_row += 1  # This ensures there is a gap between datasets

# Save the workbook
wb.save(output_excel)

print(f"Merged Excel file has been saved as '{output_excel}'")
